import argparse
from pathlib import Path
from typing import List, Tuple
from typing import Literal

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from pydantic.error_wrappers import ValidationError

import models
import profiles
import pyshacl
from __init__ import __version__


RDF_FILE_ENDINGS = ["ttl", "rdf", "xml", "json-ld", "json", "nt", "n3"]
EXCEL_FILE_ENDINGS = ["xlsx"]
KNOWN_FILE_ENDINGS = RDF_FILE_ENDINGS + EXCEL_FILE_ENDINGS


class ConversionError(Exception):
    pass


def split_and_tidy(cell_value: str):
    return [x.strip() for x in cell_value.strip().split(",")] if cell_value is not None else None


def extract_concepts_and_collections(s: Worksheet) -> Tuple[List[models.Concept], List[models.Collection]]:
    concepts = []
    collections = []
    process_concept = False
    process_collection = False
    for col in s.iter_cols(max_col=1):
        for cell in col:
            row = cell.row
            if cell.value == "Concept URI":
                process_concept = True
            elif cell.value == "Collection URI":
                process_concept = False
                process_collection = True
            elif process_concept:
                if cell.value is None:
                    pass
                else:
                    try:
                        c = models.Concept(
                            uri=s[f"A{row}"].value,
                            pref_label=s[f"B{row}"].value,
                            alt_labels=split_and_tidy(s[f"C{row}"].value),
                            definition=s[f"D{row}"].value,
                            children=split_and_tidy(s[f"E{row}"].value),
                            other_ids=split_and_tidy(s[f"F{row}"].value),
                            home_vocab_uri=s[f"G{row}"].value,
                            provenance=s[f"H{row}"].value
                        )
                        concepts.append(c)
                    except ValidationError as e:
                        raise ConversionError(f"Concept processing error, row {row}, error: {e}")
            elif process_collection:
                if cell.value is None:
                    pass
                else:
                    try:
                        c = models.Collection(
                            uri=s[f"A{row}"].value,
                            pref_label=s[f"B{row}"].value,
                            definition=s[f"C{row}"].value,
                            members=split_and_tidy(s[f"D{row}"].value),
                            provenance=s[f"E{row}"].value
                        )
                        collections.append(c)
                    except ValidationError as e:
                        raise ConversionError(f"Collection processing error, row {row}, error: {e}")
            elif cell.value is None:
                pass

    return concepts, collections


def excel_to_rdf(
        file_to_convert_path: Path,
        sheet_name=None,
        output_type: Literal["file", "string", "graph"] = "file",
        output_file_path=None
):
    """Converts a sheet within an Excel workbook to an RDF file"""
    if not file_to_convert_path.name.endswith(tuple(EXCEL_FILE_ENDINGS)):
        raise ValueError(
            "Files for conversion to RDF must be Excel files ending .xlsx"
        )
    wb = load_workbook(filename=str(file_to_convert_path), data_only=True)
    sheet = wb["vocabulary" if sheet_name is None else sheet_name]

    # Vocabulary
    try:
        cs = models.ConceptScheme(
            uri=sheet["B1"].value,
            title=sheet["B2"].value,
            description=sheet["B3"].value,
            created=sheet["B4"].value,
            modified=sheet["B5"].value,
            creator=sheet["B6"].value,
            publisher=sheet["B7"].value,
            version=sheet["B8"].value,
            provenance=sheet["B9"].value,
            custodian=sheet["B10"].value,
            pid=sheet["B11"].value,
        )
    except ValidationError as e:
        raise ConversionError(f"ConceptScheme processing error: {e}")

    # Concepts & Collections
    concepts, collections = extract_concepts_and_collections(sheet)

    # Build the total vocab
    v = models.Vocabulary(concept_scheme=cs, concepts=concepts, collections=collections)

    # Write out the file
    if output_type == "graph":
        return v.to_graph()
    elif output_type == "string":
        return v.to_graph().serialize()
    else:  # output_format == "file":
        if output_file_path is not None:
            dest = output_file_path
        else:
            dest = file_to_convert_path.with_suffix(".ttl")
        v.to_graph().serialize(destination=str(dest))
        return dest


def rdf_to_excel(
        file_to_convert_path: Path,
        profile="vocpub",
        output_file_path=None
):
    if not file_to_convert_path.name.endswith(tuple(RDF_FILE_ENDINGS)):
        raise ValueError(
            "Files for conversion to Excel must end with one of the RDF file formats: '{}'"
                .format("', '".join(RDF_FILE_ENDINGS))
        )
    if profile not in profiles.PROFILES.keys():
        raise ValueError(
            "The profile chosen for conversion must be one of '{}'. 'vocpub' is default"
                .format("', '".join(profiles.PROFILES.keys()))
        )

    # validate the RDF file
    r = pyshacl.validate(str(file_to_convert_path), shacl_graph=str(Path(__file__).parent / "validator.vocpub.ttl"))
    if not r[0]:
        raise ConversionError(
            f"The file you supplied is not valid according to the {profile} profile. The validation errors are:\n\n"
            f"{r[2]}"
        )

    # the RDF is valid so extract data and create Excel

    # wb = Workbook()
    # ws1 = wb.active
    # ws1.title = "range names"
    # for row in range(1, 40):
    #     ws1.append(range(600))
    #
    # ws2 = wb.create_sheet(title="Pi")
    # ws2['F5'] = 3.14
    # ws3 = wb.create_sheet(title="Data")
    # for row in range(10, 20):
    #     for col in range(27, 54):
    #         _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
    #
    # if output_file_path is not None:
    #     dest = output_file_path
    # else:
    #     dest = file_to_convert_path.with_suffix(".xlsx")
    # wb.save(filename=dest)
    # return dest


def main(args=None):
    parser = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter)

    parser.add_argument(
        "-v",
        "--version",
        help="The version of this copy of VocExel. Must still set an file_to_convert value to call this (can be fake)",
        action="store_true"
    )

    parser.add_argument(
        "-lp",
        "--listprofiles",
        help="This flag, if set, must be the only flag supplied. It will cause the program to list all the vocabulary"
        " profiles that this converter, indicating both their URI and their short token for use with the"
        " -p (--profile) flag when converting Excel files",
        action="store_true",
    )

    parser.add_argument(
        "file_to_convert",
        type=Path,
        help="The Excel file to convert to a SKOS vocabulary in RDF or an RDF file to convert to an Excel file",
    )

    parser.add_argument(
        "-val",
        "--validate",
        help="Validate output file",
        action="store_true"
    )

    parser.add_argument(
        "-p",
        "--profile",
        help="A profile - a specified information model - for a vocabulary. This tool understands several profiles and"
             "you can choose which one you want to convert the Excel file according to. The list of profiles - URIs "
             "and their corresponding tokens - supported by VocExcel, can be found by running the program with the "
             "flag -lp or --listprofiles.",
        default="vocpub",
    )

    parser.add_argument(
        "-ot",
        "--outputtype",
        help="The format of the vocabulary output.",
        choices=["file", "string"],
        default="file",
    )

    parser.add_argument(
        "-o",
        "--outputfile",
        help="An optionally-provided output file path.",
        required=False
    )

    parser.add_argument(
        "-s",
        "--sheet",
        help="The sheet within the target Excel Workbook to process",
        default="vocabulary",
    )

    args = parser.parse_args()

    print(args.outputfile)

    if args.listprofiles:
        s = "Profiles\nToken\tIRI\n-----\t-----\n"
        for k, v in profiles.PROFILES.items():
            s += f"{k}\t{v.uri}\n"

        print(s.rstrip())
        exit()
    elif args.version:
        print(__version__)
        exit()
    elif args.file_to_convert:
        if not args.file_to_convert.name.endswith(tuple(KNOWN_FILE_ENDINGS)):
            print("Files for conversion must either end with .xlsx (Excel) or one of the known RDF file endings, '{}'"
                  .format("', '".join(RDF_FILE_ENDINGS)))
            exit()

        print(f"Processing file {args.file_to_convert}")

        if args.file_to_convert.name.endswith(tuple(EXCEL_FILE_ENDINGS)):
            try:
                o = excel_to_rdf(args.file_to_convert, sheet_name=args.sheet, output_type=args.outputtype, output_file_path=args.outputfile)
                if args.outputtype == "string":
                    print(o)
                else:
                    print(f"Output is file {o}")
            except Exception as e:
                print(e)
                exit()
        else:  # RDF file ending
            try:
                o = rdf_to_excel(args.file_to_convert, profile=args.profile, output_file_path=args.outputfile)
                if args.outputtype == "string":
                    print(o)
                else:
                    print(f"Output is file {o}")
            except Exception as e:
                print(e)
                exit()


if __name__ == "__main__":
    import sys

    main(sys.argv[1:])
