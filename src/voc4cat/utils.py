import glob
import logging
import os
from pathlib import Path

from openpyxl import load_workbook as _load_workbook
from openpyxl.workbook.workbook import Workbook

from voc4cat.checks import Voc4catError

logger = logging.getLogger(__name__)

EXCEL_FILE_ENDINGS = [".xlsx"]
RDF_FILE_ENDINGS = {
    ".ttl": "ttl",
    ".rdf": "xml",
    ".xml": "xml",
    ".json-ld": "json-ld",
    ".json": "json-ld",
    ".nt": "nt",
    ".n3": "n3",
}
KNOWN_FILE_ENDINGS = [str(x) for x in RDF_FILE_ENDINGS] + EXCEL_FILE_ENDINGS
KNOWN_TEMPLATE_VERSIONS = ["0.4.3"]
LATEST_TEMPLATE = KNOWN_TEMPLATE_VERSIONS[-1]


class ConversionError(Exception):
    pass


def load_workbook(file_path: Path) -> Workbook:
    if file_path.suffix.lower() not in EXCEL_FILE_ENDINGS:
        msg = "Files for conversion to RDF must be xlsx files."
        raise Voc4catError(msg)
    return _load_workbook(filename=str(file_path), data_only=True)


def load_template(file_path: Path) -> Workbook:
    if file_path.suffix.lower() not in EXCEL_FILE_ENDINGS:
        msg = "Template files for RDF-to-xlsx conversion must be xlsx files."
        raise Voc4catError(msg)
    if get_template_version(load_workbook(file_path)) != LATEST_TEMPLATE:
        msg = f"Template files for RDF-to-xlsx conversion must be of latest version ({LATEST_TEMPLATE})"
        raise Voc4catError(msg)
    return _load_workbook(filename=str(file_path), data_only=True)


def get_template_version(wb: Workbook) -> str:
    # try 0.4.3 location
    try:
        intro_sheet = wb["Introduction"]
    except KeyError as exc:  # non-existing worksheet
        msg = "The version of the Excel template cannot be determined."
        logger.exception(msg)
        raise Voc4catError(msg) from exc
    return intro_sheet["J11"].value


def is_supported_template(wb):
    """Check if the template version is supported."""
    template_version = get_template_version(wb)
    if template_version not in KNOWN_TEMPLATE_VERSIONS:
        msg = f"Unsupported template version. Supported are {', '.join(KNOWN_TEMPLATE_VERSIONS)}, you supplied {template_version}."
        raise Voc4catError(msg)
    return True


def split_and_tidy(cell_value: str):
    # note this may not work in list of things that contain commas. Need to consider revising
    # to allow comma-separated values where it'll split in commas but not in things enclosed in quotes.
    if cell_value == "" or cell_value is None:
        return []
    entries = [x.strip() for x in cell_value.strip().split(",")]
    return [x for x in entries if x]


def has_file_in_multiple_formats(dir_):
    files = [
        os.path.normcase(f)
        for f in glob.glob(os.path.join(dir_, "*.*"))
        if f.endswith(tuple(KNOWN_FILE_ENDINGS))
    ]
    file_names = [os.path.splitext(f)[0] for f in files]
    unique_file_names = set(file_names)
    if len(file_names) == len(unique_file_names):
        return False
    seen = set()
    return [x for x in file_names if x in seen or seen.add(x)]
